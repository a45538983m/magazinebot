from aiogram import Router, types, F
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, FSInputFile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
from datetime import datetime, timedelta

from database.engine import get_session
from database.models import Product, Sale, Debtor, Payment

router = Router()


# ===============================================
# Состояния
# ===============================================

class ReportStates(StatesGroup):
    waiting_for_start_date = State()
    waiting_for_end_date = State()


# ===============================================
# Клавиатуры
# ===============================================

def reports_menu_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📊 Остатки (Excel)")],
            [KeyboardButton(text="📈 Продажи сегодня"), KeyboardButton(text="📈 Продажи вчера")],
            [KeyboardButton(text="📈 Продажи за месяц"), KeyboardButton(text="📅 Продажи за период")],
            [KeyboardButton(text="💰 Выручка сегодня"), KeyboardButton(text="💰 Выручка за месяц")],
            [KeyboardButton(text="🔙 Главное меню")],
        ],
        resize_keyboard=True,
    )


def cancel_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="❌ Отмена")]],
        resize_keyboard=True,
    )


# ===============================================
# ГЛАВНОЕ МЕНЮ ОТЧЁТОВ
# ===============================================

@router.message(F.text == "📥 Выгрузить")
async def reports_menu(message: types.Message):
    await message.answer(
        "📥 <b>Выгрузка отчётов</b>\n\n"
        "Выберите отчёт:",
        parse_mode="HTML",
        reply_markup=reports_menu_keyboard(),
    )


# ===============================================
# ЭКСПОРТ ОСТАТКОВ В EXCEL
# ===============================================

@router.message(F.text == "📊 Остатки (Excel)")
async def export_stock_excel(message: types.Message):
    session = get_session()
    try:
        products = session.query(Product).order_by(Product.name).all()

        if not products:
            await message.answer("📦 Нет товаров для выгрузки.")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Остатки"

        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center_align = Alignment(horizontal="center")

        headers = ["№", "Артикул", "Название", "Бренд", "Закуп. цена", "Розн. цена", "Остаток", "Место", "Подходит для"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        for i, p in enumerate(products, 2):
            compat = ""
            if p.compatible_models:
                models = [f"{cm.brand} {cm.model}" for cm in p.compatible_models[:5]]
                compat = ", ".join(models)
                if len(p.compatible_models) > 5:
                    compat += f" и ещё {len(p.compatible_models) - 5}"

            row_data = [
                i - 1, p.article, p.name, p.brand or "",
                p.purchase_price, p.selling_price, p.stock_quantity,
                p.location_code or "", compat,
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=i, column=col, value=value)
                cell.border = thin_border
                if col in [5, 6]:
                    cell.number_format = "#,##0.00"
                elif col == 7:
                    cell.alignment = center_align

            if p.stock_quantity == 0:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=i, column=col).fill = red_fill
            elif p.stock_quantity < 10:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=i, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

        widths = [5, 18, 35, 15, 12, 12, 10, 10, 40]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

        filename = f"stocks_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = f"temp/{filename}"
        os.makedirs("temp", exist_ok=True)
        wb.save(filepath)

        file = FSInputFile(filepath)
        await message.answer_document(
            file,
            caption=f"📊 Остатки товаров на {datetime.now().strftime('%d.%m.%Y %H:%M')}",
        )

        os.remove(filepath)

    finally:
        session.close()


# ===============================================
# ПРОДАЖИ ЗА СЕГОДНЯ
# ===============================================

@router.message(F.text == "📈 Продажи сегодня")
async def sales_today(message: types.Message):
    today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    await export_sales_excel(message, today_start, "сегодня")


# ===============================================
# ПРОДАЖИ ЗА ВЧЕРА
# ===============================================

@router.message(F.text == "📈 Продажи вчера")
async def sales_yesterday(message: types.Message):
    yesterday_start = (datetime.now() - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
    yesterday_end = yesterday_start.replace(hour=23, minute=59, second=59)
    await export_sales_excel(message, yesterday_start, "вчера", yesterday_end)


# ===============================================
# ПРОДАЖИ ЗА МЕСЯЦ
# ===============================================

@router.message(F.text == "📈 Продажи за месяц")
async def sales_month(message: types.Message):
    month_start = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    await export_sales_excel(message, month_start, "за месяц")


# ===============================================
# ПРОДАЖИ ЗА ПЕРИОД
# ===============================================

@router.message(F.text == "📅 Продажи за период")
async def sales_period_start(message: types.Message, state: FSMContext):
    await state.set_state(ReportStates.waiting_for_start_date)
    await message.answer(
        "📅 <b>Продажи за период</b>\n\n"
        "Введите начальную дату в формате <b>ДД.ММ.ГГГГ</b> (например: 01.01.2024):",
        parse_mode="HTML",
        reply_markup=cancel_keyboard(),
    )


@router.message(ReportStates.waiting_for_start_date)
async def process_start_date(message: types.Message, state: FSMContext):
    try:
        start_date = datetime.strptime(message.text.strip(), "%d.%m.%Y")
    except ValueError:
        await message.answer("⚠️ Неверный формат даты. Введите ДД.ММ.ГГГГ:")
        return

    await state.update_data(start_date=start_date)
    await state.set_state(ReportStates.waiting_for_end_date)
    await message.answer(
        "Введите конечную дату в формате <b>ДД.ММ.ГГГГ</b>:",
        parse_mode="HTML",
        reply_markup=cancel_keyboard(),
    )


@router.message(ReportStates.waiting_for_end_date)
async def process_end_date(message: types.Message, state: FSMContext):
    try:
        end_date = datetime.strptime(message.text.strip(), "%d.%m.%Y")
        end_date = end_date.replace(hour=23, minute=59, second=59)
    except ValueError:
        await message.answer("⚠️ Неверный формат даты. Введите ДД.ММ.ГГГГ:")
        return

    data = await state.get_data()
    start_date = data["start_date"]

    if end_date < start_date:
        await message.answer("⚠️ Конечная дата не может быть раньше начальной. Введите заново:")
        return

    await state.clear()
    await export_sales_excel(message, start_date, f"с {start_date.strftime('%d.%m.%Y')} по {end_date.strftime('%d.%m.%Y')}", end_date)


# ===============================================
# ФОРМИРОВАНИЕ EXCEL С ПРОДАЖАМИ
# ===============================================

async def export_sales_excel(message, start_date, period_name, end_date=None):
    if end_date is None:
        end_date = datetime.now()

    session = get_session()
    try:
        sales = (
            session.query(Sale)
            .filter(Sale.created_at >= start_date, Sale.created_at <= end_date)
            .order_by(Sale.created_at.desc())
            .all()
        )

        if not sales:
            await message.answer(f"📈 Продаж {period_name} не было.")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Продажи"

        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center_align = Alignment(horizontal="center")

        headers = ["№", "Дата", "Товар", "Артикул", "Кол-во", "Цена за ед.", "Сумма", "Должник", "Оплачено", "Долг"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        total_sum = 0
        total_paid = 0
        total_debt = 0

        for i, s in enumerate(sales, 2):
            product_name = s.product.name if s.product else "Товар удалён"
            product_article = s.product.article if s.product else "-"
            debtor_name = s.debtor.name if s.debtor else ""
            paid_amount = sum(p.amount for p in s.payments) if s.debtor_id else s.total_amount
            debt_amount = s.total_amount - paid_amount if s.debtor_id else 0

            row_data = [
                i - 1, s.created_at.strftime('%d.%m.%Y %H:%M'),
                product_name, product_article,
                s.quantity, s.price, s.total_amount,
                debtor_name, paid_amount, debt_amount,
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=i, column=col, value=value)
                cell.border = thin_border
                if col in [6, 7, 9, 10]:
                    cell.number_format = "#,##0.00"

            total_sum += s.total_amount
            total_paid += paid_amount
            total_debt += debt_amount

        summary_row = len(sales) + 2
        ws.cell(row=summary_row, column=6, value="ИТОГО:").font = Font(bold=True)
        ws.cell(row=summary_row, column=7, value=total_sum).font = Font(bold=True)
        ws.cell(row=summary_row, column=7).number_format = "#,##0.00"
        ws.cell(row=summary_row, column=9, value=total_paid).font = Font(bold=True)
        ws.cell(row=summary_row, column=9).number_format = "#,##0.00"
        ws.cell(row=summary_row, column=10, value=total_debt).font = Font(bold=True)
        ws.cell(row=summary_row, column=10).number_format = "#,##0.00"

        for col in range(1, len(headers) + 1):
            ws.cell(row=summary_row, column=col).border = thin_border

        widths = [5, 18, 30, 18, 8, 12, 12, 15, 12, 10]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

        filename = f"sales_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = f"temp/{filename}"
        os.makedirs("temp", exist_ok=True)
        wb.save(filepath)

        file = FSInputFile(filepath)

        text = f"📈 <b>Продажи {period_name}</b>\n\n"
        text += f"Количество продаж: <b>{len(sales)}</b>\n"
        text += f"Общая сумма: <b>{total_sum:.2f}</b>\n"
        text += f"Оплачено: <b>{total_paid:.2f}</b>\n"
        if total_debt > 0:
            text += f"Осталось долгов: <b>{total_debt:.2f}</b>\n"

        await message.answer(text, parse_mode="HTML")
        await message.answer_document(file, caption=f"📋 Детальный отчёт {period_name}")

        os.remove(filepath)

    finally:
        session.close()


# ===============================================
# ВЫРУЧКА СЕГОДНЯ
# ===============================================

@router.message(F.text == "💰 Выручка сегодня")
async def revenue_today(message: types.Message):
    today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    await show_revenue(message, today_start, "сегодня")


# ===============================================
# ВЫРУЧКА ЗА МЕСЯЦ
# ===============================================

@router.message(F.text == "💰 Выручка за месяц")
async def revenue_month(message: types.Message):
    month_start = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    await show_revenue(message, month_start, "за месяц")


# ===============================================
# РАСЧЁТ ВЫРУЧКИ
# ===============================================

async def show_revenue(message, start_date, period_name):
    session = get_session()
    try:
        sales = (
            session.query(Sale)
            .filter(Sale.created_at >= start_date)
            .order_by(Sale.created_at.desc())
            .all()
        )

        if not sales:
            await message.answer(f"💰 Продаж {period_name} не было.")
            return

        total_sales = len(sales)
        total_items = sum(s.quantity for s in sales)
        total_revenue = sum(s.total_amount for s in sales)

        # Оплаченные и долги
        paid_amount = 0
        debt_amount = 0
        for s in sales:
            if s.debtor_id:
                paid = sum(p.amount for p in s.payments)
                paid_amount += paid
                debt_amount += (s.total_amount - paid)
            else:
                paid_amount += s.total_amount

        # Прибыль (выручка - закупочная стоимость)
        total_profit = 0
        for s in sales:
            if s.product:
                profit_per_item = s.price - s.product.purchase_price
                total_profit += profit_per_item * s.quantity

        # Топ-5 товаров по продажам
        from sqlalchemy import func as sqlfunc
        top_products = (
            session.query(Product.name, sqlfunc.sum(Sale.quantity).label("qty"))
            .join(Sale, Sale.product_id == Product.id)
            .filter(Sale.created_at >= start_date)
            .group_by(Product.id)
            .order_by(sqlfunc.sum(Sale.quantity).desc())
            .limit(5)
            .all()
        )

        text = (
            f"💰 <b>ВЫРУЧКА {period_name.upper()}</b>\n\n"
            f"📦 Количество продаж: <b>{total_sales}</b>\n"
            f"📊 Продано товаров: <b>{total_items} шт.</b>\n"
            f"💵 Выручка: <b>{total_revenue:.2f}</b>\n"
            f"✅ Оплачено: <b>{paid_amount:.2f}</b>\n"
        )

        if debt_amount > 0:
            text += f"📝 В долгу: <b>{debt_amount:.2f}</b>\n"

        text += f"📈 Прибыль: <b>{total_profit:.2f}</b>\n"

        if top_products:
            text += f"\n🏆 <b>Топ-5 товаров:</b>\n"
            for i, (name, qty) in enumerate(top_products, 1):
                text += f"  {i}. {name} — {qty} шт.\n"

        await message.answer(text, parse_mode="HTML")

    finally:
        session.close()


# ===============================================
# ОТМЕНА
# ===============================================

@router.message(F.text == "❌ Отмена")
async def cancel_report(message: types.Message, state: FSMContext):
    current_state = await state.get_state()
    if current_state is not None:
        await state.clear()
    await message.answer(
        "📥 <b>Выгрузка отчётов</b>\n\nВыберите отчёт:",
        parse_mode="HTML",
        reply_markup=reports_menu_keyboard(),
    )